#!/usr/bin/env python

import os
import re
import sys
import logging
import argparse
import datetime
import json
import io
import zipfile
import itertools
import functools

import openpyxl
import openpyxl.utils
import openpyxl.styles
import openpyxl.styles.colors
import openpyxl.writer.excel
from openpyxl.comments import Comment

import requests

import neil_tools
from neil_tools import spreadsheet_tools
from neil_tools import gen_templates

import config as config_static
import web_session
import message
import veh_stats

import arc_o365
#from O365_local.excel import WorkBook as o365_WorkBook
import O365
from O365.excel import WorkBook as o365_WorkBook



NOW = datetime.datetime.now().astimezone()
DATESTAMP = NOW.strftime("%Y-%m-%d")
TIMESTAMP = NOW.strftime("%Y-%m-%d %H:%M:%S %Z")
FILESTAMP = NOW.strftime("%Y-%m-%d %H-%M-%S %Z")
EMAILSTAMP = NOW.strftime("%Y-%m-%d %H-%M")

# flag field in vehicle structures
IN_AVIS = '__IN_AVIS__'

# flag field in avis object array
AVIS_SOURCE = '__AVIS_SOURCE__'
AVIS_SOURCE_OPEN       = 'OPEN'
AVIS_SOURCE_OPEN_ALL   = 'OPEN_ALL'
AVIS_SOURCE_CLOSED     = 'CLOSED'
AVIS_SOURCE_CLOSED_ALL = 'CLOSED_ALL'
AVIS_SOURCE_MISSING    = 'MISSING'

FILL_RED = openpyxl.styles.PatternFill(fgColor="FFC0C0", fill_type = "solid")
FILL_GREEN = openpyxl.styles.PatternFill(fgColor="C0FFC0", fill_type = "solid")
FILL_YELLOW = openpyxl.styles.PatternFill(fgColor="FFFFC0", fill_type = "solid")
FILL_BLUE = openpyxl.styles.PatternFill(fgColor="5BB1CD", fill_type = "solid")
FILL_CYAN = openpyxl.styles.PatternFill(fgColor="A0FFFF", fill_type = "solid")

# flag in person object that they have a vehicle
PERSON_VEHICLES = '__PERSON_VEHICLES__'
PERSON_REF = '__PERSON_REF__'
ROSTER_REF = '__ROSTER_REF__'
DISTRICT = '__DISTRICT__'
TnM = '__TnM__'

COMMENT_AUTHOR = "Avis Report Reconciler Program"

NO_GAP_GROUP = 'ZZZ-No-GAP'
MASTER = 'Master'


def main():
    args = parse_args()
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
    log.debug("running...")

    config = neil_tools.init_config(config_static, ".env")

    errors = False
    for dr in args.dr_id:
        if dr not in config.DR_CONFIGURATIONS:
            log.error(f"specified DR ID '{ dr }' not found in configured DRs.  Valid DRs are { list(config.DR_CONFIGURATIONS.keys()) }")
            errors = True

    if errors:
        sys.exit(1)


    account_avis = None
    for dr in args.dr_id:
        roster = None
        dr_config = config.DR_CONFIGURATIONS[dr]

        # fetch from DTT
        session = web_session.get_session(config, dr_config)
        success = get_dr_list(config, dr_config, session)
        if not success:
            log.info(f"Login failure for dr { dr }: retrying without cookies")
            session = web_session.get_session(config, dr_config, force_new_session=True)
            success = get_dr_list(config, dr_config, session)
            if not success:
                log.error(f"Could not access DTT for dr { dr }")
                errors = True
                continue

        log.debug(f"DR_ID { dr }")

        # get people and vehicles from the DTT
        vehicles = get_vehicles(config, dr_config, args, session)
        people = get_people(config, dr_config, args, session)
        agencies = get_agencies(config, dr_config, args, session)

        account_mail = None

        if args.send or args.test_send:
            log.debug(f"initializing mail account: { dr_config.token_filename }")
            account_mail = init_o365(config, dr_config.token_filename)


        if args.do_car or args.do_no_car:
            if not roster:
                roster = get_roster(config, dr, vehicles, people)

            do_status_messages(dr_config, args, account_mail, vehicles, people, roster)


        # avis report
        if args.do_avis:
            if account_avis == None:
                account_avis = init_o365(config, config.TOKEN_FILENAME_AVIS)

            # fetch the avis spreadsheet
            output_bytes = make_avis(config, dr_config, account_avis, vehicles, agencies)

            file_name = f"DR{ dr_config.dr_id } { FILESTAMP } Avis Report.xlsx"
            if args.store:
                store_report(config, account_avis, file_name, output_bytes)

            # save a local copy
            log.debug(f"storing avis report to { file_name }")
            with open(file_name, "wb") as fb:
                fb.write(output_bytes)

            if args.send or args.test_send:
                send_avis_report(dr_config, args, account_mail, file_name)


            if not args.save:
                os.remove(file_name)

        # group vehicle report
        if args.do_group:
            if not roster:
                roster = get_roster(config, dr, vehicles, people)

            # generate the group report
            output_bytes = make_group_report(config, dr_config, args, vehicles, people, roster)
            file_name = f"DR{ dr_config.dr_id } { FILESTAMP } Group Vehicle Report.xlsx"

            if args.store:
                store_report(config, account_mail, file_name, output_bytes)

            # save a local copy for attachment
            log.debug(f"storing gap report to { file_name }")
            with open(file_name, "wb") as fb:
                fb.write(output_bytes)

            if args.send or args.test_send:
                send_group_report(dr_config, args, account_mail, file_name)

            if not args.save:
                os.remove(file_name)

        if args.do_vehicles:
            output_bytes = make_vehicle_backup(config, dr_config, vehicles)

            if args.store:
                item_name = f"DR{ dr_config.dr_id } { FILESTAMP } Vehicle Backup.xlsx"
                store_report(config, account_mail, item_name, output_bytes)

        if args.do_dtr:
            dtr = make_dtr(dr_config, vehicles)

    if errors:
        sys.exit(1)


def get_roster(config, dr, vehicles, people):
    roster_contents = message.fetch_dr_roster(config, dr)
    roster = message.convert_roster_to_objects(roster_contents)

    key_name = 'Mem#'

    # construct a dict of roster entries, keyed by VC member id; convert member_id index from float to string
    roster_by_vc = dict(
            map(lambda d: (str(int(d[key_name])), d),
                filter(lambda d: key_name in d, roster)
                )
            )

    preprocess_people_roster(vehicles, people, roster_by_vc)
    return roster_by_vc


def get_dr_list(config, dr_config, session):
    """ Get the list of DRs we have access to from the DTT """

    url = config.DTT_URL + "Vehicles"

    r = session.get(url)
    r.raise_for_status()

    codes = r.html.find('#DisasterCodes', first=True)
    if codes == None:
        # could not read form: login failed?
        return None

    options = codes.find('option')

    for option in options:
        value = option.attrs['value']
        text = option.text
        log.debug(f"option value { value } text { text }")

    # for now, while we only have access to one DR: pick the last value
    dr_config.id = value
    dr_config.name = text

    return value



def make_vehicle_backup(config, dr_config, vehicles):
    """ generate a spreadsheet of all the vehicles as a backup """

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Backup")

    class ColumnDef:

        def __init__(name, dtype="string", width=10):
            self._name = name
            self._dtype = dtype
            self._width = width

        @property 
        def name(self):
            return self._name

        @property 
        def dtype(self):
            return self._dtype

        @property 
        def width(self):
            return self._width

    columns = [
            ColumnDef("Expiring"),
            ColumnDef("Ctg"),
            ]

    for row in vehicles:
        for v in row['Vehicle']:
            pass


    wb_buffer = workbook_to_buffer(wb)
    return wb_buffer


def send_avis_report(dr_config, args, account, file_name):
    """ send an email with the avis report (contained in file_name).

        Respect the args.send and args.test_send flags (at least one of which must be set
    """

    message_body = \
f"""
<p>
Hello everyone.  This is an automated report matching vehicles in the DTT against
a list of vehicles provided by Avis.  The Avis list tends to be delayed by 24-48 hours.
</p>

<p>
If you have suggestions about the contents of the report or bugs in the program itself,
or have other tasks you think should be automated on a DR: email
<a href='mailto:{ dr_config.program_email }'>{ dr_config.program_email }</a>.
</p>
"""

    send_report_common(dr_config, args, account, file_name, "Avis Report", message_body, dr_config.reply_email)



def send_group_report(dr_config, args, account, file_name):
    """ send an email with the group vehicle report (contained in file_name).

        Respect the args.send and args.test_send flags (at least one of which must be set
    """

    message_body = \
f"""
<p>
Hello everyone.  This is an automated report showing vehicles on the DR organized by GAP 'Group'
</p>

<p>
If you have any updates or corrections: please send them to
<a href='mailto:{ dr_config.reply_email }'>{ dr_config.reply_email }</a>.
</p>

<p>
If you have suggestions about the contents of the report or bugs in the program itself,
or have other tasks you think should be automated on a DR: email
<a href='mailto:{ dr_config.program_email }'>{ dr_config.program_email }</a>.
</p>
"""

    send_report_common(dr_config, args, account, file_name, "Group Vehicle Report", message_body, dr_config.target_list)



def send_report_common(dr_config, args, account, file_name, report_type, message_body, dest_email):

    mailbox = account.mailbox()
    message = mailbox.new_message(resource=dr_config.send_mail)

    if args.test_send:
        message.bcc.add(dr_config.email_bcc)

    if args.send:
        message.bcc.add(dest_email)
        log.debug(f"sending { file_name } to { dest_email }")
        posting = f"<p>This message was sent to { dest_email }.  Please do *not* reply to the whole list</p>\n"
    else:
        log.debug(f"not sending { file_name } to { dest_email }")
        posting = \
f"""
<p>
DEBUG Version: not sent to the list
</p>
"""

    message.body = \
f"""
<!DOCTYPE html>
<html>
<meta http-equiv="Content-type" content="text/html" charset="UTF8" />
<title>DR{ dr_config.dr_id } { report_type }</title>
</head>
<body>

<h1>DR{ dr_config.dr_id } { report_type }</h1>
{ posting }

{ message_body }

</body>
</html>
"""

    message.subject = file_name
    message.attachments.add( file_name )

    try:
        message.send(save_to_sent_folder=True)
    except requests.RequestException as e:
        log.error(f"got an error: { e }, response json { e.response.json }")
        raise e



def fetch_avis(config, account):
    """ get the most recent avis workbook """

    storage = account.storage()
    drive = storage.get_drive(config.NHQDCSDLC_DRIVEID)
    fy21 = drive.get_item_by_path(config.FY21_ITEM_PATH)

    children = fy21.get_items()

    rental_re = re.compile(r'^ARC Open Rentals\s*-?\s*(\d{1,2})-(\d{1,2})-(\d{2,4})\.xlsx$')
    count = 0
    mismatch = 0
    newest_file_date = None
    newest_ref = None
    for child in children:

        #log.debug(f"child { child.name } id { child.object_id }")

        # humans made the name, so there is good (but not perfect) adherence to a naming standard.
        # its something like "ARC Open Rentals - mm-dd-yyyy.xlsx", but there is variation...
        match = rental_re.match(child.name)
        count += 1
        if match is None:
            log.info(f"no pattern match for '{ child.name }'")
            mismatch += 1
        else:
            month = match.group(1).lstrip('0')
            day = match.group(2).lstrip('0')
            year = match.group(3).lstrip('0')

            #log.debug(f"file match: { year }-{ month }-{ day }")

            file_date = datetime.date(int(year), int(month), int(day))
            if newest_file_date is None or newest_file_date < file_date:
                newest_file_date = file_date
                newest_ref = child

    log.debug(f"total items: { count } mismatched { mismatch }")

    if newest_ref is None:
        raise Exception("make_avis: no valid files found")
    log.debug(f"newest_file { newest_ref.name }")
    config['AVIS_FILE'] = newest_ref.name

    workbook = o365_WorkBook(newest_ref, persist=False)

    return workbook

def make_avis(config, dr_config, account, vehicles, agencies):
    """ fetch the latest avis vehicle report from sharepoint """

    avis_wb = fetch_avis(config, account)

    output_wb = openpyxl.Workbook()

    # insert vehicles
    output_ws_open = output_wb.create_sheet(f"DR{ dr_config.dr_id }")
    log.debug(f"sheet names: { output_wb.sheetnames }")

    # we now have the latest file.  Suck out all the data
    avis_open_title, avis_open_columns, avis_open, avis_open_all = read_avis_sheet(dr_config, avis_wb.get_worksheet('Open RA'))
    add_missing_avis_vehicles(vehicles, avis_open_all, avis_open, closed=False)

    #avis_closed_title, avis_closed_columns, avis_closed, avis_closed_all = read_avis_sheet(dr_config, avis_wb.get_worksheet('Closed RA'))
    #add_missing_avis_vehicles(vehicles, avis_closed_all, avis_closed, closed=True)

    # generate the 'Open RA' sheet
    output_columns = copy_avis_sheet(output_ws_open, avis_open_columns, avis_open_title, avis_open)
    match_avis_sheet(output_ws_open, output_columns, avis_open, vehicles, agencies)

    # insert overview at the end
    insert_avis_overview(output_wb, config, dr_config)

    # now serialize the workbook
    bufferview = workbook_to_buffer(output_wb)

    return bufferview



def workbook_to_buffer(wb):
    """ serialize a workbook to a byte array so it can be saved (either to network or locally) """

    # cleanup: delete the default sheet name in the output workbook
    default_sheet_name = 'Sheet'
    if default_sheet_name in wb:
        del wb[default_sheet_name]


    iobuffer = io.BytesIO()
    zipbuffer = zipfile.ZipFile(iobuffer, mode='w')
    writer = openpyxl.writer.excel.ExcelWriter(wb, zipbuffer)
    writer.save()

    bufferview = iobuffer.getbuffer()
    log.debug(f"output spreadsheet length: { len(bufferview) }")

    return bufferview



def copy_avis_sheet(ws, columns, title, rows):
    """ copy avis data to the output ws """
    wrap_alignment = openpyxl.styles.Alignment(wrapText=True, horizontal='center')

    fixed_width_columns = {
            'Rental Region Desc': 20,
            'Rental Distict Desc': 20,
            'CO Date': 16,
            'Rental Loc Desc': 22,
            'Address Line 1': 30,
            'Address Line 3': 30,
            'Full Name': 30,
            'Exp CI Date': 16,
            'AWD Orgn Buildup Desc': 30,
            }

    # first write out the title
    row = 1
    col = 1
    output_columns = {}
    for index, key in enumerate(title):
        if key != '' and key != 'CO Time' and key != 'Exp CI Time' and not key.startswith('__'):
            cell = ws.cell(row=row, column=col, value=key)
            cell.alignment = wrap_alignment
            col_letter = openpyxl.utils.get_column_letter(col)

            if key in fixed_width_columns:
                ws.column_dimensions[col_letter].width = fixed_width_columns[key]
            else:
                ws.column_dimensions[col_letter].auto_size = True

            output_columns[key] = col
            col += 1

    # now add the data
    time_regex = re.compile(r'(\d{2}):(\d{2}):(\d{2})')
    row = 1
    for row_data in rows:
        row += 1
        col = 1
        for key in output_columns:


            # ignore columns without a name and meta-entries
            if key == '' or key.startswith('__'):
                continue

            if key not in row_data:
                col += 1
                continue

            value = row_data[key]

            # fold time columns into their date columns
            time_key = None
            if key == 'CO Date':
                time_key = 'CO Time'
            elif key == 'Exp CI Date':
                time_key = 'Exp CI Time'

            # process date/time column pairs
            if time_key:

                if isinstance(value, datetime.datetime):
                    dt = value
                else:
                    dt = spreadsheet_tools.excel_to_dt(value)

                    time_string = row_data[time_key]
                    time_match = time_regex.match(time_string)
                    if time_match:
                        interval = datetime.timedelta(hours=int(time_match.group(1)), minutes=int(time_match.group(2)), seconds=int(time_match.group(3)))
                        dt += interval
                    else:
                        log.debug(f"copy_avis_sheet: row { row } time { time_key } didn't parse: '{ time_string }'")

                #log.debug(f"adding row { row } column { col } title { key } value { dt } time_string '{ time_string }'")
                cell = ws.cell(row=row, column=col, value=dt)
                cell.number_format = 'yyyy-mm-dd hh:mm'

            # don't output time columns: they were already handled
            elif key == 'CO Time' or key == 'Exp CI Time':
                # don't increment column counter and ignore time columns
                continue

            # copy over all other columns
            else:

                #log.debug(f"adding row { row } column { col } title { key } value { value }")
                ws.cell(row=row, column=col, value=value)

            col += 1

    last_col_letter = openpyxl.utils.get_column_letter(col -1)
    table_ref = f"A1:{last_col_letter}{row}"
    log.debug(f"last col letter { last_col_letter }, table_ref { table_ref }")
    table = openpyxl.worksheet.table.Table(displayName='AvisOpen', ref=table_ref)
    ws.add_table(table)

    return output_columns



def cleanup_v_field(vehicle, field_name):
    """ do field cleanup for vehicle entries """

    if field_name not in vehicle:
        return None

    value = vehicle[field_name]

    if value is None:
        return None

    value = value.strip()

    if field_name == 'RentalAgreementNumber':
        # rental agreements always start with 'U'; add it if not present
        if value[0] != 'U':
            value = 'U' + value

    elif field_name == 'RentalAgreementReservationNumber':
        # the emailed reservation number looks like 12345678-US-6; put it in cannonical form
        value = value.replace('-', '').upper()

    elif field_name == 'KeyNumber':
        # make sure there are 9 digits in the number
        value = value.rjust(9, '0')


    return value


def make_vehicle_index(vehicles, first_field, second_field=None, reservation=False, keynumber=False):

    #log.debug(f"make_vehicle_index: vehicles len { len(vehicles) }, 1st_field { first_field } 2nd_field { second_field }")
    result = {}
    if first_field == 'DisasterVehicleID':
        for row in vehicles:
            key = row[first_field]
            if key in result:
                log.error(f"Duplicate key { key } in vehicles: old { result[key] } new { row }")
            else:
                result[key] = row

    else:
        for row in vehicles:
            vehicle = row['Vehicle']

            if first_field not in vehicle:
                continue

            if second_field is not None and second_field not in vehicle:
                continue

            key = cleanup_v_field(vehicle, first_field)
            if key is None:
                continue

            if second_field is not None:
                second = vehicle[second_field]
                if second is None:
                    continue
                key = key + " " + second.strip()

            key = key.upper()
            #log.debug(f"make_vehicle_index: key '{ key }'")

            if key in result:
                old_row = result[key]

                if old_row['Status'] == 'Active' and row['Status'] == 'Active':
                    log.error(f"Error: Duplicate rows with key '{ key }' and both are active")
                    log.info(f"DUPLICATE: key '{ key }' is already in result.\n\nExisting{ result[key] }\n\nnew { row }")

                else:
                    # at most one row is active
                    log.info(f"Error: Duplicate rows with key '{ key }', at most one active "
                            f"(old: '{ old_row['Vehicle']['KeyNumber'] }', "
                            f"new '{ row['Vehicle']['KeyNumber'] }')"
                            )

                    if row['Status'] == 'Active':
                        # we know that old_row is not active, so use the current row.
                        result[key] = row
            else:
                # first row with this key
                result[key] = row

    return result

def add_missing_avis_vehicles(vehicles, avis_all, avis_open, closed):
    """ find Avis vehicles that are not in the Avis report """

    # generate some indexes to look up values faster
    i_ra  =   spreadsheet_tools.make_index(avis_all, 'Rental Agreement No')
    i_res =   spreadsheet_tools.make_index(avis_all, 'Reservation No')
    i_key =   spreadsheet_tools.make_index(avis_all, 'MVA No')
    i_plate = spreadsheet_tools.make_index(avis_all, 'License Plate State Code', 'License Plate Number')

    # note: using avis_open for this index
    i_row =   spreadsheet_tools.make_index(avis_open, spreadsheet_tools.ROW_INDEX)

    # this is duplicating work in match_avis_sheet, but it doesn't seem worth it to save the info
    v_ra = make_vehicle_index(vehicles, 'RentalAgreementNumber')
    v_res= make_vehicle_index(vehicles, 'RentalAgreementReservationNumber')
    v_key = make_vehicle_index(vehicles, 'KeyNumber')
    v_plate = make_vehicle_index(vehicles, 'PlateState', 'Plate')

    missing = []


    # walk through the avis_open sheet and record all the matches
    for row in avis_open:

        ra = row['Rental Agreement No']
        res = row['Reservation No']
        key = row['MVA No']
        plate = row['License Plate State Code'] + ' ' + row['License Plate Number']
        addr_line_3 = row['Address Line 3']

        # use DisasterVehicleID as the DTT identity for a vehicle
        ra_id = get_dtt_id(v_ra, ra)
        res_id = get_dtt_id(v_res, res)
        key_id = get_dtt_id(v_key, key)
        plate_id = get_dtt_id(v_plate, plate)
        
        row[AVIS_SOURCE] = AVIS_SOURCE_OPEN


    # get_dtt_id() records which vehicles were looked up.  Add vehicles in DTT that are not in AVIS
    for record in vehicles:
        # vehicle is not active
        if record['Status'] != 'Active':
            continue

        # already in the Avis report
        if IN_AVIS in record:
            continue

        # make sure its an Avis rental
        if record['Vehicle']['Vendor'] != 'Avis':
            continue

        # see if this vehicle is in avis_all
        vehicle = record['Vehicle']

        if vehicle['KeyNumber'] is None:
            # don't bother if there is no key number: its just a reservation
            continue

        ra = cleanup_v_field(vehicle, 'RentalAgreementNumber')
        res = cleanup_v_field(vehicle, 'RentalAgreementReservationNumber')
        key = cleanup_v_field(vehicle, 'KeyNumber')
        plate = f"{ vehicle['PlateState'] } { vehicle['Plate'].strip() }"

        if vehicle['Plate'] == "812KBK":
            log.debug(f"saw vehicle 812KBK")

        if ra in i_ra or res in i_res or key in i_key or plate in i_plate:

            #log.debug(f"Adding missing vehicle to OPEN { key }")

            # one or more fields from the DTT appeared in the avis_all list, but not the avis_open list.
            # add the row to the avis_open  list
            new_avis_records = []
            if ra in i_ra:
                new_avis_records.append(i_ra[ra])
            if res in i_res:
                new_avis_records.append(i_res[res])
            if key in i_key:
                new_avis_records.append(i_key[key])
            if plate in i_plate:
                new_avis_records.append(i_plate[plate])

            # add any records to avis_open that we don't already have
            for row in new_avis_records:
                row_index = row[spreadsheet_tools.ROW_INDEX]
                if row_index not in i_row:
                    # this row isn't in avis_open yet...
                    row[AVIS_SOURCE] = AVIS_SOURCE_OPEN_ALL
                    avis_open.append(row)
                    i_row[row_index] = row

        else:
            # ZZZ: need to check closed roster before adding missing vehicles...
            if True:
                # this is an entirely new vehicle that doesn't match anything in avis_all
                #log.debug(f"Adding missing vehicle to ALL { key }")

                row = make_avis_from_vehicle(record)
                row[AVIS_SOURCE] = AVIS_SOURCE_MISSING
                row[spreadsheet_tools.ROW_INDEX] = len(avis_all) + 1
                missing.append(row)


    # add the missing rows to the end
    avis_open.extend(missing)
    avis_all.extend(missing)


def make_avis_from_vehicle(record):
    """ synthesize an entirely new avis record from a vehicle record """

    vehicle = record['Vehicle']

    def get_field(name):
        """ utility function to safely get field values """
        value = cleanup_v_field(vehicle, name)
        if value is None:
            return ""
        return value

    avis = {}

    # map avis fields to vehicle fields
    fields = {
            'MVA No':                   'KeyNumber',
            'License Plate State Code': 'PlateState',
            'License Plate Number':     'Plate',
            'Make':                     'Make',
            'Model':                    'Model',
            'Ext Color Code':           'Color',
            'Reservation No':           'RentalAgreementReservationNumber',
            'Rental Agreement No':      'RentalAgreementNumber',
            'Full Name':                'RentalAgreementPerson',
            'Rental Loc Desc':          'PickupAgencyName',
            }

    avis['Cost Control No'] = 'MISSING'
    for f_avis, f_vehicle in fields.items():
        avis[f_avis] = get_field(f_vehicle)

    # strip off the timezone; openpyxl can't write it
    pickupDate = re.sub(r'-\d\d:\d\d$', '', get_field('RentalAgreementPickupDate'))
    avis['CO Date'] = datetime.datetime.fromisoformat(pickupDate)

    #log.debug(f"new avis record: { avis }")
    return avis



def get_dtt_id(vehicle_dict, index):
    """ utility function to look up a field in the vehicle object from the DTT """

    index = index.upper()
    if index not in vehicle_dict:
        return None

    # mark the vehicle as found-in-avis-report
    vehicle_dict[index][IN_AVIS] = True

    #log.debug(f"index { index } v_dict { vehicle_dict[index] }")
    return vehicle_dict[index]['DisasterVehicleID']



STRIKE_FONT = openpyxl.styles.Font(strike=True)
def mark_cell(ws, fill, v_id, vid, row_num, col_map, col_name, comment=None):
    """ apply a fill to a particular cell

        v_id is a dictionary of vids to rows
        vid is the vehicle id number
    """

    col_num = col_map[col_name]

    cell = ws.cell(row=row_num, column=col_num)

    if fill is not None:
        cell.fill = fill

    if comment is not  None:
        cell.comment = comment

    if v_id is not None and vid is not None:
        status = v_id[vid]['Status']
        if status != 'Active':
            log.debug(f"inactive vehicle found: { vid } status '{ status }'")
            cell.font = STRIKE_FONT


agency_key_fixup_punctuation = re.compile(r'[.,]')
agency_key_fixup_space = re.compile(r'  +')
def make_agency_key(agency):
    """ generate a (hopefully) unique key for the DTT agencies.  Combine Address, City, State, Zip.

        Name seems the obvious choice, but DTT names don't match AVIS names at all
    """

    address = agency['Address']
    city = agency['City']
    state = agency['State']
    zipcode = agency['Zip']

    key = f"{ address }/{ city } { state }{ zipcode }".upper()
    #log.debug(f"key { key }")

    # remove punctuation, compress spaces
    key = agency_key_fixup_punctuation.sub('', key)
    key = agency_key_fixup_space.sub(' ', key)
    return key

dtt_to_avis_make_dict = {
        'Buick': 'BUIC',
        'Chevrolet': 'CHEV',
        'Chrysler': 'CHRY',
        'Dodge': 'DODG',
        'Ford': 'FORD',
        'GMC': 'GMC ',
        'Honda': 'HOND',
        'Hyundai': 'HYUN',
        'Jeep': 'JEEP',
        'Kia': 'KIA ',
        'Mazda': 'MAZD',
        'Mitsubishi': 'MITS',
        'Nissan': 'NISS',
        'Subaru': 'SUBA',
        'Toyota': 'TOYO',
        'Volkswagen': 'VOLK',
        }
dtt_to_avis_model_dict = {
        '3': '3SED',
        '4Runner': '4RUN',
        '6': '6SED',
        'Acadia': 'ACA2',
        'Accord': 'ACCO',
        'Altima': 'ALTI',
        'Cadenza': 'K5K5',
        'Camry': 'CAMR',
        'Caravan': 'GRCA',
        'Charger': 'CHRT',
        'Civic': 'CIVI',
        'CJ': 'GLH4',
        'Compass': 'CMPS',
        'Corolla': 'CRLA',
        'CR-V': 'CRV4',
        'CX-5': 'CX5A',
        'CX-9': 'CX9F',
        'Durango': 'DURA',
        'Eclipse': 'ECCF',
        'Ecosport': 'ECOA',
        'Edge': 'EDE2',
        'Elantra': 'ELAN',
        'Envision': 'ENVI',
        'Equinox': 'EQUI',
        'Escape': 'ESCA',
        'Escort': 'ECOA',
        'Expedition': 'EXL4',
        'Explorer': 'EXL2',
        'F-150': 'F150',
        'Forester': 'FORE',
        'Forte': 'FORT',
        'Frontier': 'FRO4',
        'Fusion': 'FUSI',
        'Golf': 'GOLF',
        'Highlander': 'HIGH',
        'HR-V': 'HRVA',
        'Jetta': 'JETT',
        'Journey': 'JOU2',
        'Legacy': 'LEGA',
        'Malibu': 'MALB',
        'Mustang': 'MUST',
        'Optima': 'OPTI',
        'Outback': 'OUTB',
        'Outlander': 'OUTL',
        'Pacifica': 'PACI',
        'Passat': 'PASS',
        'Pathfinder': 'PATH',
        'Prius': 'PRIH',
        'RAM': 'RAR2',
        'Ranger': 'RAN4',
        'RAV 4': 'RAV4',
        'Rogue': 'ROG2',
        'Santa Fe': 'SANT',
        'Sedona': 'SEDO',
        'Sentra': 'SENT',
        'Sienna': 'SIEN',
        'Sonata': 'SONA',
        'Sorento': 'SO7F',
        'Soul': 'SOUL',
        'Sportage': 'SPO2',
        'Tacoma': 'TAC4',
        'Tahoe': 'TAHO',
        'Terrain': 'TERR',
        'Tiguan': 'TIG2',
        'Tracker': 'TRX2',
        'Traverse': 'TRAV',
        'Tucson': 'TUCS',
        'Versa': 'VRSA',       
        'Voyager': 'VGER',
        'Wrangler': 'WRA4',
        'XV Crosstrex': 'XVCR',
        'Yukon': 'YUS4',
        }
dtt_to_avis_color_dict = {
        'Silver': 'SIL',
        'White': 'WHI',
        'Blue': 'BLU',
        'Black': 'BLK',
        'Gray': 'GRY',
        'Red': 'RED',
        'Brown': 'BRO',
        }

def dtt_to_avis_make(make_tuple, avis_tuple):
    """ convert DTT make/model/color to avis """

    make = None
    model  = None
    color = None

    if make_tuple[0] in dtt_to_avis_make_dict:
        make = dtt_to_avis_make_dict[make_tuple[0]]
    else:
        log.debug(f"could not find make '{ make_tuple[0] }' / avis '{ avis_tuple[0] }'")
        
    if make_tuple[1] in dtt_to_avis_model_dict:
        model = dtt_to_avis_model_dict[make_tuple[1]]
    else:
        log.debug(f"could not find model '{ make_tuple[1] }' / avis '{ avis_tuple[1] }'")

    if make_tuple[2] in dtt_to_avis_color_dict:
        color = dtt_to_avis_color_dict[make_tuple[2]]
    else:
        # avis does't have Yellow and Beige in its list of colors
        if make_tuple[2] != 'Yellow' and make_tuple[2] != 'Beige':
            log.debug(f"could not find color '{ make_tuple[2] }' / avis '{ avis_tuple[2] }'")


    return (make, model, color)

def match_avis_sheet(ws, columns, avis, vehicles, agencies):
    """ match entries from the DTT to entries in the Avis report. """

    # generate different index for the vehicles
    v_ra = make_vehicle_index(vehicles, 'RentalAgreementNumber')
    v_res= make_vehicle_index(vehicles, 'RentalAgreementReservationNumber')
    v_key = make_vehicle_index(vehicles, 'KeyNumber')
    v_plate = make_vehicle_index(vehicles, 'PlateState', 'Plate')
    v_id = make_vehicle_index(vehicles, 'DisasterVehicleID')

    vid_dict = dict( (v['DisasterVehicleID'], v) for v in vehicles)

    #log.debug(f"plate keys: { v_plate.keys() }")

    multispace_re = re.compile(r'\s+')

    #log.debug(f"make translation dict: { dtt_to_avis_model_dict }")
    agency_no_match_list = {}

    def mark_cell_wrapper(vid, field_name, spreadsheet_row, columns, column_name):
        if vid is None:
            fill = FILL_RED
            comment = None
        else:
            fill = FILL_YELLOW
            vrow = vid_dict[vid]
            veh = vrow['Vehicle']
            veh_value = veh[field_name]

            comment = Comment(
                    f"DTT id { vid } -- status { vrow['Status'] }\n"
                    f"Driver { veh['CurrentDriverName'] }\n"
                    f"Key { veh['KeyNumber'] }\n"
                    f"Reservation { veh['RentalAgreementReservationNumber'] }\n"
                    f"Agreement { veh['RentalAgreementNumber'] }\n"
                    f"Plate { veh['PlateState'] } { veh['Plate'] }\n"
                    , COMMENT_AUTHOR, height=300, width=400)

        mark_cell(ws, fill, v_id, vid, spreadsheet_row, columns, column_name, comment=comment)

    spreadsheet_row = 1
    for row in avis:
        spreadsheet_row += 1

        ra = row['Rental Agreement No']
        res = row['Reservation No']
        key = row['MVA No']
        plate = row['License Plate State Code'] + ' ' + row['License Plate Number']
        cost_control = row['Cost Control No']

        if plate == "WA BOS5423":
            log.debug(f"saw avis plate { plate }")

        if cost_control == 'MISSING':
            # this is a synthetic row created from the DTT, not AVIS; don't bother matching
            continue

        addr_line = ''
        if 'Address Line 1' in row and 'Address Line 3' in row:
            addr_line = f"{ row['Address Line 1'] }/{ row['Address Line 3'] }"
            addr_line = multispace_re.sub(' ', addr_line)

        # use DisasterVehicleID as the DTT identity for a vehicle
        ra_id = get_dtt_id(v_ra, ra)
        res_id = get_dtt_id(v_res, res)
        key_id = get_dtt_id(v_key, key)
        plate_id = get_dtt_id(v_plate, plate)

        #log.debug(f"ra { ra_id } res { res_id } key { key_id } plate { plate_id }; raw { ra } { res } { key } { plate }")

        if AVIS_SOURCE in row:
            avis_source = row[AVIS_SOURCE]
        else:
            avis_source = None

        if ra_id == None and res_id == None and key_id == None and plate_id == None:
            # vehicle doesn't appear in the DTT at all; color it blue
            fill = FILL_BLUE

            if avis_source == AVIS_SOURCE_OPEN_ALL:
                fill = FILL_CYAN

            mark_cell(ws, fill, v_id, ra_id,    spreadsheet_row, columns, 'Rental Agreement No')
            mark_cell(ws, fill, v_id, res_id,   spreadsheet_row, columns, 'Reservation No')
            mark_cell(ws, fill, v_id, key_id,   spreadsheet_row, columns, 'MVA No')
            mark_cell(ws, fill, v_id, plate_id, spreadsheet_row, columns, 'License Plate State Code')
            mark_cell(ws, fill, v_id, plate_id, spreadsheet_row, columns, 'License Plate Number')

        elif ra_id == res_id and ra_id == key_id and ra_id == plate_id:
            # all columns match: color it green

            fill = FILL_GREEN

            #if avis_source == AVIS_SOURCE_OPEN_ALL:
            #    fill = FILL_CYAN

            mark_cell(ws, fill, v_id, ra_id,    spreadsheet_row, columns, 'Rental Agreement No')
            mark_cell(ws, fill, v_id, res_id,   spreadsheet_row, columns, 'Reservation No')
            mark_cell(ws, fill, v_id, key_id,   spreadsheet_row, columns, 'MVA No')
            mark_cell(ws, fill, v_id, plate_id, spreadsheet_row, columns, 'License Plate State Code')
            mark_cell(ws, fill, v_id, plate_id, spreadsheet_row, columns, 'License Plate Number')

            # since all four 'unique' fields match: check additional fields
            vrow = vid_dict[ra_id]
            vehicle = vrow['Vehicle']

            # check 'agency' (aka pickup location)
            agency_key = vehicle['PickupAgencyId']
            if agency_key not in agencies:
                log.debug(f"could not find agency key '{ agency_key }' in v_agencies ")
            else:
                agency = agencies[agency_key]
                agency_string = agency['AvisAgencyString']
                comment = None

                if agency_string == addr_line:
                    fill = FILL_GREEN
                else:
                    fill = FILL_YELLOW
                    if agency_string not in agency_no_match_list:
                        # only print once per agency
                        log.debug(f"no agency match '{ addr_line }' / '{ agency_string }'")
                        agency_no_match_list[agency_string] = True

                    comment = Comment(
                            f"DTT location is:\n"
                            f"{ agency['Name'] }\n"
                            f"{ agency['Address'] }\n"
                            f"{ agency['City'] } { agency['State'] }{ agency['Zip'] }",
                            COMMENT_AUTHOR, height=300, width=400)

                mark_cell(ws, fill, None, None, spreadsheet_row, columns, 'Rental Loc Desc', comment=comment)
                mark_cell(ws, fill, None, None, spreadsheet_row, columns, 'Address Line 1')
                mark_cell(ws, fill, None, None, spreadsheet_row, columns, 'Address Line 3')


            # check pickup and expected dropoff date
            for (dtt_col, avis_col) in (('RentalAgreementPickupDate', 'CO Date'), ('DueDate', 'Exp CI Date')):
                dtt_date = datetime.datetime.fromisoformat(vehicle[dtt_col]).date()
                col_num = columns[avis_col]
                cell = ws.cell(row=spreadsheet_row, column=col_num)

                avis_pickup_dt = cell.value

                if avis_pickup_dt != None:

                    avis_pickup_date = avis_pickup_dt.date()

                    if dtt_date == avis_pickup_date:
                        fill = FILL_GREEN
                    else:
                        #log.debug(f"pickup date: key { key } dtt { dtt_date }/{vehicle[dtt_col]} avis { avis_pickup_date }/{ avis_pickup_dt }")
                        fill = FILL_YELLOW
                        cell.comment = Comment(f"DTT date is { dtt_date }", COMMENT_AUTHOR, height=300, width=400)

                    cell.fill = fill

            # check make/model/color
            avis_make_cols = ['Make', 'Model', 'Ext Color Code']
            avis_veh_make = list(row[col_name] for col_name in avis_make_cols)
            #avis_veh_make = (row['Make'], row['Model'], row['Ext Color Code'])
            dtt_veh_make_orig = (vehicle['Make'].strip() if vehicle['Make'] != None else None,
                    vehicle['Model'].strip() if vehicle['Model'] != None else None,
                    vehicle['Color'].strip() if vehicle['Color'] != None else None)
            dtt_veh_make = dtt_to_avis_make(dtt_veh_make_orig, avis_veh_make)

            #log.debug(f"avis make {  avis_veh_make } dtt { dtt_veh_make } orig { dtt_veh_make_orig }")

            for (i, col_name) in enumerate(avis_make_cols):
                v = dtt_veh_make[i]

                fill = None
                comment = None

                if v is None:
                    # DTT string not found in our mapping table; ignore
                    comment = Comment(f"No mapping for DTT value { dtt_veh_make_orig[i] }", COMMENT_AUTHOR, height=300, width=400)

                else:

                    if v == avis_veh_make[i]:
                        fill = FILL_GREEN
                    else:
                        fill = FILL_YELLOW
                        comment = Comment(f"DTT value is { dtt_veh_make_orig[i] } -> { v }", COMMENT_AUTHOR, height=300, width=400)

                mark_cell(ws, fill, None, None, spreadsheet_row, columns, col_name, comment=comment)
            

        else:
            # else color yellow if value is found; red if value not found
            mark_cell_wrapper(ra_id, 'RentalAgreementNumber', spreadsheet_row, columns, 'Rental Agreement No')
            mark_cell_wrapper(res_id, 'RentalAgreementReservationNumber', spreadsheet_row, columns, 'Reservation No')
            mark_cell_wrapper(key_id, 'KeyNumber', spreadsheet_row, columns, 'MVA No')
            mark_cell_wrapper(plate_id, 'Plate', spreadsheet_row, columns, 'License Plate Number')

            mark_cell(ws, FILL_RED if plate_id is None else FILL_YELLOW, v_id, plate_id, spreadsheet_row, columns, 'License Plate State Code')

        if avis_source == AVIS_SOURCE_OPEN_ALL:
            mark_cell(ws, FILL_YELLOW, None, None, spreadsheet_row, columns, 'Cost Control No')
        elif avis_source == AVIS_SOURCE_MISSING or avis_source is None:
            mark_cell(ws, FILL_CYAN, None, None, spreadsheet_row, columns, 'Cost Control No')






        





def insert_avis_overview(wb, config, dr_config):
    """ insert an overview (documentation) sheet in the workbook """

    doc_string = f"""
This document has vehicles from the daily Avis report for this DR ({ dr_config.dr_id }).

On the Open RA (rental agreement) sheet there should be one row per vehicle marked as assigned to the DR
(from the 'Cost Control No' column).

Rows in the Open RA sheet have been checked against the DTT Vehicles data.  These columns are checked between
the two reports: License Plate State Code/License Plate Number, Reservation No, Rental Agreement No, Reservation No.

If all four fields match: the cells will be marked Green.

If a vehicle in the Avis report is not found in the DTT: the cells will be blue.

If all four fields don't match: any field that is not found in the DTT will be red.
Otherwise the fields will be yellow.  Red fields are usually typos in the DTT data entry.
A row of all yellow often means a data entry error where fields from different vehicles were
entered on the same DTT entry.

Cyan fields are used when a vehicle is in the DTT but not in the Avis report.

If the MVA, Reservation, Contract, or Plate fields use strike-through fonts: the vehicle is released
in the DTT but active in the Avis report.

The 'Cost Control No' column is a bit different.  This column is used to encode the DR number.
The program tries to figure out if which DR matches, but it is not perfect.

If there is a vehicle in the DTT for this DR that has a different entry in the Cost Control No
field: the entry will be yellow.

This file based on the { config.AVIS_FILE } file
This file generated at { TIMESTAMP }

"""

    ws = insert_overview(wb, doc_string)

    cell = ws.cell(row=2, column=1, value="Example Green cell")
    cell.fill = FILL_GREEN
    cell = ws.cell(row=3, column=1, value="Example Yellow cell")
    cell.fill = FILL_YELLOW
    cell = ws.cell(row=4, column=1, value="Example Red cell")
    cell.fill = FILL_RED
    cell = ws.cell(row=5, column=1, value="Example Blue cell")
    cell.fill = FILL_BLUE
    cell = ws.cell(row=6, column=1, value="Example Cyan cell")
    cell.fill = FILL_CYAN

    return ws

def insert_group_overview(wb, dr_config):

    doc_string = f"""
This document has vehicles for DR{ dr_config.dr_id }.

Asset Management is the responsibility of the Group Leads.  This report shows
who has a vehicle that Transportation has in its database.

Please help us keep the database up to date.  Send all updates to { dr_config.reply_email }.

Notes:

* Groups are separated into separate tabs.

This file generated at { TIMESTAMP }

"""

    ws = insert_overview(wb, doc_string)
    return ws




def insert_overview(wb, doc_string):

    ws = wb.create_sheet('Overview')

    ws.column_dimensions['A'].width = 120       # hard coded width....
    ws.row_dimensions[1].height = 550           # hard coded height....

    cell = ws.cell(row=1, column=1, value=doc_string)
    cell.alignment = openpyxl.styles.Alignment(wrapText=True, vertical='top')

    return ws

def read_avis_sheet(dr_config, sheet):

    log.debug(f"sheet name { sheet.name }")
    sheet_range = sheet.get_used_range()

    #log.debug(f"last_column { sheet_range.get_last_column() } last_row { sheet_range.get_last_row() }")

    values = sheet_range.values

    values.pop(0)               # sheet title text -- before the 'data table'

    #title_row = values.pop(0)   # column headers
    #log.debug(f"title_row { title_row }")

    title_row = values[0]
    avis_columns = spreadsheet_tools.title_to_dict(title_row)
    avis_all = spreadsheet_tools.matrix_to_object_array(values)

    # the DR number format (in the 'Cost Control No' column) isn't well 
    dr_column = 'Cost Control No'

    # trying to match patterns like:
    # 98, 098, DR098, DR098-21, DR098-2021, 098-21, etc...
    avis_dr = []
    for dr_tuple in dr_config.get_dr_list():
        dr_num = dr_tuple[0]
        dr_year = dr_tuple[1]

        dr_num = dr_num.lstrip('0')
        dr_regex = re.compile(r"(dr)?\s*0*" f"{ dr_num }" r"(-(20)?" f"{ dr_year })?", flags=re.IGNORECASE)

        f_result = filter(lambda row: dr_regex.match(row[dr_column]), avis_all)
        avis_dr.extend( f_result )

    log.debug(f"found { len(avis_dr) } vehicles associated with the DR")
    
    return title_row, avis_columns, avis_dr, avis_all




def init_o365(config, token_filename=None):
    """ do initial setup to get a handle on office 365 graph api """

    if token_filename != None:
        o365 = arc_o365.arc_o365.arc_o365(config, token_filename=token_filename)
    else:
        o365 = arc_o365.arc_o365.arc_o365(config)

    account = o365.get_account()
    if account is None:
        raise Exception("could not access office 365 graph api")

    return account

    

def get_people(config, dr_config, args, session):
    """ Retrieve the people list from the DTT (as a json list) """

    data = get_json(config, dr_config, args, session, 'People/Details')

    # construct a dict keyed by PersonID
    d = dict( (d['PersonID'], d) for d in data )
    return d


def get_vehicles(config, dr_config, args, session):
    """ Retrieve the vehicle list from the DTT (as a json list) """

    data = get_json(config, dr_config, args, session, 'Vehicles')
    return data

def get_agencies(config, dr_config, args, session):
    """ retrieve the current rental agencies for this DR

        return a dict keyed by the AgencyID
    """

    data = get_json(config, dr_config, args, session, 'Agencies', prefix='api/Disasters/')

    # construct a dict keyed by AgencyID from the data array
    d = dict( (h['AgencyID'], h) for h in data )

    for h in data:
        h['AvisAgencyString'] = make_agency_key(h)



    #log.debug(f"got agencies: { d }")
    return d



def get_json(config, dr_config, args, session, api_type, prefix='api/Disaster/'):

    file_name = f"cached_{ dr_config.dr_id }_{ re.sub(r'/.*', '', api_type) }.json"

    data = None
    if args.cached_input:
        log.debug(f"reading cached input from { file_name }")
        with open(file_name, "rb") as f:
            buffer = f.read()

        data = json.loads(buffer)

    else:

        url = config.DTT_URL + f"{ prefix }{ dr_config.id }/" + api_type

        r = session.get(url)
        r.raise_for_status()

        #log.debug(f"response headers { r.headers }")
        #log.debug(f"response { r.content }")

        if args.save_input:
            log.debug(f"saving to { file_name }")
            with open(file_name, "wb") as f:
                f.write(r.content)

        data = r.json()
        #log.debug(f"r.status { r.status_code } r.reason { r.reason } r.url { r.url } r.content_type { r.headers['content-type'] } data rows { len(data) }")

    #log.debug(f"json { data }")
    #log.debug(f"Returned data\n{ json.dumps(data, indent=2, sort_keys=True) }")

    return data


def store_report(config, account, item_name, wb_bytes):
    """ store the report.  for now: stash it in the Transportation/Reports file store. """

    log.debug("called")
    storage = account.storage()
    drive = storage.get_drive(config.TRANS_REPORTS_DRIVEID)
    reports = drive.get_item_by_path(config.TRANS_REPORTS_FOLDER_PATH)

    stream = io.BytesIO(wb_bytes)
    stream.seek(0, io.SEEK_END)
    stream_size = stream.tell()
    stream.seek(0, io.SEEK_SET)

    log.debug(f"stream_size is { stream_size }")

    result = reports.upload_file(item_name, item_name=item_name, stream=stream, stream_size=stream_size, conflict_handling='replace')
    log.debug(f"after upload: result { result }")


def make_group_report(config, dr_config, args, vehicles, people, roster):
    """ make a workbook of all vehicles, arranged by GAPs """

    wb = openpyxl.Workbook()
    insert_group_overview(wb, dr_config)

    # sort the vehicles by GAP
    vehicles = sorted(vehicles, key=lambda x: vehicle_to_gap(x['Vehicle']))

    # add the master sheet
    add_gap_sheet(wb, MASTER, vehicles, people, roster, activity=False)

    gap_list, district_list = get_vehicle_gap_groups(vehicles)

    #log.debug(f"district_list: { district_list }")

    for group in gap_list:
        group_vehicles = filter_by_gap_group(group, vehicles)
        #log.debug(f"make_group_report: group { group } # veh { len(list(group_vehicles)) }")
        add_gap_sheet(wb, group, group_vehicles, people, roster, activity=True)

    # serialize the workbook
    bufferview = workbook_to_buffer(wb)

    return bufferview


gap_to_group_re = re.compile('^([A-Z]+)/')
gap_to_activity_re = re.compile('^([A-Z]+/[A-Z]*)/')
motor_pool_re = re.compile(r'^Motor Pool \(([^\)]*)\)$')
def vehicle_to_group(vehicle, activity=False):
    """ turn a Gap (Group/Activity/Position) name into just the Group portion """

    gap = vehicle['GAP']
    if gap is None:
        gap = ''

    driver = vehicle['CurrentDriverName']

    pool = vehicle_in_pool(vehicle)
    if pool is not None:
        return pool

    if activity:
        group = gap_to_activity_re.match(gap)
    else:
        group = gap_to_group_re.match(gap)

    if group is not None:
        group_name = group.group(1)
        #log.debug(f"Found a group '{ group_name }'")
        return group_name

    return NO_GAP_GROUP


def vehicle_to_gap(vehicle):
    """ returns the motor pool name if its a pool, otherwise None """

    pool = vehicle_in_pool(vehicle)
    if pool is not None:
        return pool

    gap = vehicle['GAP']
    if gap is None:
        gap = ''

    return gap



def vehicle_in_pool(vehicle):
    """ return pool name if vehicle is in a pool, otherwise None """
    driver = vehicle['CurrentDriverName']

    if driver is None:
        return None

    pool = motor_pool_re.match(driver)

    if pool is not None:
        pool_name = pool.group(1)
        #log.debug(f"Found a pool '{ pool_name }'")
        return pool_name

    return None



def get_vehicle_gap_groups(vehicles):
    """ return a sorted list of groups from the gaps in the vehicles """

    group_re = re.compile('^([A-Z]+)')

    groups = {}
    districts = {}

    for row in vehicles:

        if row['Status'] != 'Active':
            continue

        vehicle = row['Vehicle']

        gap = vehicle['GAP']
        group = vehicle_to_group(vehicle)

        if group is None:
            if gap != '':
                log.info(f"Count not find Group in gap '{ gap }' vehicle { vehicle }")
            else:
                groups[NO_GAP_GROUP] = True
                pass
            pass
        else:
            # remember that we saw this group; don't worry about duplicates
            #log.debug(f"Adding group { group } from gap { gap }")
            groups[group] = True

        if DISTRICT in vehicle:
            district = vehicle[DISTRICT]
            districts[district] = True


    group_list = sorted(groups.keys())
    #log.debug(f"group_list { group_list }")
    return group_list, list(districts.keys())



def filter_by_district(district, vehicles):
    """ only return vehicles in the given district """

    results = []
    for row in vehicles:
        veh = row['Vehicle']

        v_district = veh.get(DISTRICT)
        if v_district == district:
            results.append(row)

    return results

def filter_by_gap_group(group, vehicles):
    """ only return the vehicles whos GAP starts with the group """

    return list(filter(lambda x: vehicle_to_group(x['Vehicle']) == group, vehicles))

def get_people_column(people, pid, column):
    # look up a person by id (from the vehicle row) and return the specified column
    if pid in people:
        person = people[pid]
        if column in person:
            return person[column]
    else:
        if pid != None:
            log.debug(f"could not find pid { pid } in people dict")

    # we didn't find the value
    return '' 


def add_gap_sheet(wb, sheet_name, vehicles, people, roster, activity=False):
    """ make a new sheet with the specified vehicles """

    #log.debug(f"processing sheet '{ sheet_name }' activity { activity }")
    if sheet_name == '':
        sheet_name = "BLANK"


    ws = wb.create_sheet(sheet_name)

    column_defs = {
            'GAP':          { 'width': 15, 'key': lambda x: vehicle_to_gap(x), },
            'Driver':       { 'width': 30, 'key': lambda x: x['CurrentDriverName'], },
            'Key Number':   { 'width': 12, 'key': lambda x: x['KeyNumber'], },
            'Vendor':       { 'width': 20, 'key': lambda x: x['Vendor'], },
            'Car Info':     { 'width': 25, 'key': lambda x: f"{ x['Make'] } { x['Model'] } { x['Color'] }", },
            'Plate':        { 'width': 15, 'key': lambda x: f"{ x['PlateState'] } { x['Plate'] }", },
            'Type':         { 'width': 10, 'key': lambda x: x['VehicleType'], },
            'District':     { 'width': 10, 'key': lambda x: x['District'], },
            'Location':     { 'width': 30, 'key': lambda x: x['CurrentDriverWorkLocationName'], },
            'Lodging':      { 'width': 30, 'key': lambda x: x['CurrentDriverLodging'], },
            'Reservation':  { 'width': 15, 'key': lambda x: x['RentalAgreementReservationNumber'], },
            'Outprocessed': { 'width': 15, 'key': lambda x: "" if not x['OutProcessed'] else "OUTPROCESSED", },
            'Cell Phone':   { 'width': 13, 'key': lambda x: get_people_column(people, x['CurrentDriverPersonId'], 'MobilePhone'), },
            'Email':        { 'width': 20, 'key': lambda x: get_people_column(people, x['CurrentDriverPersonId'], 'Email'), },
        }

    row = 1
    col = 0
    for key, defs in column_defs.items():
        col = col + 1
        title = key
        cell = ws.cell(row=row, column=col, value=title)

        col_letter = openpyxl.utils.get_column_letter(col)
        if 'width' in defs:
            #log.debug(f"setting column { col_letter } to width { defs['width'] }")
            ws.column_dimensions[col_letter].width = defs['width']
        else:
            ws.column_dimensions[col_letter].auto_size = True

    ws.freeze_panes = 'C2'

    seen_groups = {}
    seen_groups['ALL' if sheet_name == MASTER else sheet_name] = True
    row = 1
    previous_group = None
    for vehicle in vehicles:

        # only process active vehicles
        if vehicle['Status'] != 'Active':
            #log.debug(f"ignorning inactive vehicle { vehicle }")
            continue

        row += 1
        col = 0

        row_vehicle = vehicle['Vehicle']
        group = vehicle_to_group(row_vehicle, activity=activity)
        if group not in seen_groups:
            #log.debug(f"seen new group: '{ group }'")
            seen_groups[group] = True

        # leave a blank row when the group changes
        if group != previous_group:
            if previous_group is not None:
                row += 1
            previous_group = group

        for key, defs in column_defs.items():
            col = col + 1

            try:
                lookup_func = defs['key']
                value = lookup_func(row_vehicle)

                if value is None or value == 'None' or value == 'None None' or value == 'None None None':
                    # no valid data; make it a blank
                    value = ''
                #log.debug(f"adding cell { row },{ col }, value { value }")
                cell = ws.cell(row=row, column=col, value=value)
            except:
                log.info(f"Could not expand column { key } row_vehicle { row_vehicle }: { sys.exc_info()[0] }")

    # ZZZ: horrible hack; for some reason the no-gap table makes excel unhappy; I haven't been
    # able to figure out why, so I'm just not adding a table for that sheet
    if sheet_name != NO_GAP_GROUP:

        # table names can't have spaces
        sheet_name = re.sub(' ', '_', sheet_name)

        last_col_letter = openpyxl.utils.get_column_letter(col)
        table_ref = f"A1:{ last_col_letter }{ row }"
        table = openpyxl.worksheet.table.Table(displayName=f"{ sheet_name }Table", ref=table_ref)

        ws.add_table(table)
    else:
        log.debug(f"Ignoring table for ZZZ-No-Gap sheet")

    # now do stats for the groups
    total_vehicles = 0
    stat_results = []
    for g in seen_groups.keys():
        roster_count, vehicle_count = veh_stats.compute_stats(vehicles, roster, g)
        total_vehicles += vehicle_count
        stat_results.append({ 'group': g, 'roster': roster_count, 'vehicle': vehicle_count })

    if total_vehicles != 0:
        row += 2
        ws.cell(row=row, column=1, value="Vehicle Statistics")
        row += 1
        ws.cell(row=row, column=1, value="GAP")
        ws.cell(row=row, column=2, value="# MDA")
        ws.cell(row=row, column=3, value="Rentals")
        ws.cell(row=row, column=4, value="Ratio").number_format = "#,###.00"
        row += 1

        for stat in stat_results:
            roster_count = stat['roster']
            vehicle_count = stat['vehicle']

            if vehicle_count == 0:
                continue

            ratio = roster_count / vehicle_count

            ws.cell(row=row, column=1, value=stat['group'])
            ws.cell(row=row, column=2, value=roster_count)
            ws.cell(row=row, column=3, value=vehicle_count)
            ws.cell(row=row, column=4, value=ratio).number_format = "#,###.00"
            row += 1


def preprocess_people_roster(vehicles, people, roster_by_vc):
    """ mark all persons who have vehicles, and link vehicle to people and roster
        entries (if they exist)
    """

    for row in vehicles:

        # ignore non-active vehicles
        if row['Status'] != 'Active':
            continue

        vehicle = row['Vehicle']

        driver_id = vehicle['CurrentDriverPersonId']
        driver = vehicle['CurrentDriverName']

        if driver_id is None:
            # no driver assigned yet
            continue

        pool = motor_pool_re.match(driver)
        if pool is not None:
            # pool vehicle; don't bother messaging
            continue


        if driver_id not in people:
            # assumption is we should always be able to find a driver
            log.error(f"Could not find driver_id { driver_id } in person list ({ vehicle['CurrentDriverNameAndMemberNo'] })")
            continue

        person = people[driver_id]
        vehicle[PERSON_REF] = person

        # mark this person as having a vehicle
        if PERSON_VEHICLES not in person:
            person[PERSON_VEHICLES] = []
        person[PERSON_VEHICLES].append(row)
        if driver_id == 6517:
            log.debug(f"appending vehicle key # { vehicle['KeyNumber'] } for driver { driver_id }")

        #log.debug(f"marking person { driver_id } / { driver } as having a vehicle")

        if ROSTER_REF not in person:
            vc = person['VC']
            roster_ref = roster_by_vc.get(vc)
            person[ROSTER_REF] = roster_ref
            vehicle[ROSTER_REF] = roster_ref
        else:
            roster_ref = person[ROSTER_REF]
            vehicle[ROSTER_REF] = roster_ref

        if roster_ref is not None:
            vehicle[DISTRICT] = roster_ref['District']
            vehicle[TnM] = roster_ref['T&M']
        else:
            vehicle[TnM] = ''



def do_status_messages(dr_config, args, account, vehicles, people, roster_by_vc):

    vehicle_code_map = {
            'A': 'ERV',
            'B': 'Box Truck',
            'C': 'Chapter Vehicle',
            'D': 'Delivery Vehicle',
            'R': 'Rental Vehicle',
            'L': 'Loaned Vehicle',
            'POV': 'Personal Vehicle',
            }
    def vehicle_category_code_to_string(code):
        #string = vehicle_code_map.get(code)
        #log.debug(f"vehicle_category: code '{ code }' string '{ string }'")
        return vehicle_code_map.get(code)

    key_name = 'Mem#'


    # make an index of person ids

    templates = gen_templates.init()
    date = datetime.datetime.now().strftime("%Y-%m-%d %H%M")

    templates = gen_templates.init()
    t_vehicle = templates.get_template("mail_vehicle.html")
    t_no_veh = templates.get_template("mail_no_veh.html")

    # now generate the emails
    if args.do_car:
        count = 0
        for person in people.values():

            if PERSON_VEHICLES not in person:
                continue

            l = person[PERSON_VEHICLES]

            first_name = person['FirstName']
            last_name = person['LastName']
            email = person['Email']

            log.debug(f"person { first_name } { last_name } has { len(l) }")

            context = {
                    'first_name': first_name,
                    'last_name': last_name,
                    'email': email,
                    'vehicles': l,
                    'reply_email': dr_config.reply_email,
                    'date': date,
                    'vehicle_category_code_to_string': vehicle_category_code_to_string,
                    }

            body = t_vehicle.render(context)

            #log.debug(f"body { body }")

            if account is not None:
                m = account.new_message(resource=dr_config.send_email)
                if args.test_send:
                    m.bcc.add(dr_config.email_bcc)
                if args.send:
                    m.to.add(email)

                m.subject = f"Vehicle Status - { date } - { first_name } { last_name }"
                m.body = body

                if args.test_send or args.send:
                    m.send()

                
            # debug only
            count += 1
            if args.mail_limit and count >= args.mail_limit:
                break


        log.debug(f"found { count } people with cars")


    # now do people without vehicles

    if args.do_no_car:
        count = 0
        for i, person in enumerate(people.values()):
            
            first_name = person['FirstName']
            last_name = person['LastName']
            email = person['Email']

            # ignore people that have vehicles
            if PERSON_VEHICLES in person:
                #log.debug(f"ignoring person { first_name } { last_name }: has a vehicle")
                continue

            status = person['Status']

            # don't hassle folks who are off the job
            if status != 'Checked In':
                #log.debug(f"ignoring person { first_name } { last_name }: status { status } not 'Checked In'")
                continue

            # checked in status is often wrong; cross check against the most recent VC roster
            vc_num = person['VC']

            if vc_num not in roster_by_vc:
                # assume they are no longer on the DR and ignore them
                #log.debug(f"ignoring person { first_name } { last_name }: not in VC roster: vc_num '{ vc_num }'")
                continue

            vc_entry = roster_by_vc[vc_num]
            loc_type = vc_entry['Location type']

            if loc_type == 'Virtual':
                #log.debug(f"ignoring person { first_name } { last_name }: location type is virtual")
                continue

            log.debug(f"({ i }) person { first_name } { last_name } { person['PersonID'] } { status } has no vehicles")

            context = {
                    'first_name': first_name,
                    'last_name': last_name,
                    'email': email,
                    'reply_email': dr_config.reply_email,
                    'date': date,
                    }

            body = t_no_veh.render(context)

            #log.debug(f"body { body }")

            if account is not None:
                m = account.new_message(resource=dr_config.send_email)
                if args.test_send:
                    m.bcc.add(dr_config.email_bcc)
                if args.send:
                    m.to.add(email)

                m.subject = f"Vehicle Status - { date } - { first_name } { last_name }"
                m.body = body

                if args.test_send or args.send:
                    m.send()

            
            # debug only
            count += 1
            if args.mail_limit and count >= args.mail_limit:
                break

        log.debug(f"found { count } people without cars")


def make_dtr(dr_config, vehicles):
    """ do a replacement for the daily transportation report.  Can be pretty stupid for now"""

    count_category_active = {}
    count_category_all = {}
    count_dvid = {}
    count_vid = {}

    count_rental_group = {}
    count_rental_psc = {}
    count_boxtruck_group = {}
    count_boxtruck_psc = {}
    count_passengervan_group = {}
    count_passengervan_psc = {}
    count_cargovan_group = {}
    count_cargovan_psc = {}

    def bump_count(count_dict, index):
        if index not in count_dict:
            count_dict[index] = 0
        count_dict[index] += 1

    gap_to_group_re = re.compile(r'([A-Z]+)/.*')
    def gap_to_group(gap):
        if gap == '' or gap is None:
            return 'LOG'

        for prefix in [ 'IDC/DMH', 'IDC/DHS', 'IDC/DSC', 'IP/DA', 'REC/REV', 'RES/DAT', 'ER/PA', 'ER/APAT', 'ER/FR' ]:
            if gap.startswith(f'{ prefix }/'):
                return prefix

        group = gap_to_group_re.sub(r'\1', gap)
        return group

    group_to_psc = {
            'MC': 21,

            'RES/DAT': 22,

            'IDC': 23,
            'IDC/DHS': 23,
            'IDC/DMH': 24,
            'IDC/DSC': 24,

            'ER': 26,
            'IP/DA': 26,

            'REC': 27,
            'REC/REV': 27,

            'OM': 28,
            'IP': 28,
            'FIN': 28,
            'ER/PA': 28,
            'ER/APAT': 28,

            'DST': 29,
            'LOG': 29,
            'SS': 29,
            'RES': 29,

            'ER/FR': 80,
            }

    # compute category sub-totals
    for row in vehicles:
        veh = row['Vehicle']

        status = row['Status']
        category = veh['VehicleCategoryCode']
        type = veh['VehicleType']
        dvid = row['DisasterVehicleID']
        vid = veh['VehicleID']
        gap = veh['GAP']
        group = gap_to_group(gap)
        psc = group_to_psc[group]
        #log.debug(f"gap '{ gap }' group '{ group }'")
        #if type != 'Car':
        #    log.debug(f"type: '{ type }'")

        if category == 'R' or category == 'D':
            if type == 'Passenger Van':
                category = 'P-Van'
            elif type == 'Cargo Van' or category == 'D':
                category = 'D'
            else:
                category = 'R'

        bump_count(count_category_all, category)
        if status == 'Active':
            bump_count(count_category_active, category)
            if category == 'B':
                bump_count(count_boxtruck_group, group)
                bump_count(count_boxtruck_psc, psc)
            elif category == 'P-Van':
                bump_count(count_passengervan_group, group)
                bump_count(count_passengervan_psc, psc)
            elif category == 'D':
                bump_count(count_cargovan_group, group)
                bump_count(count_cargovan_psc, psc)
            elif category == 'R':
                bump_count(count_rental_group, group)
                bump_count(count_rental_psc, psc)

        bump_count(count_dvid, dvid)
        bump_count(count_vid, vid)

    def sumcategory(count_dict, keys):
        total = 0
        for key in keys:
            total += count_dict.get(key, 0)

        #log.debug(f"sumcategory: total '{ total }'")
        return total


    # this is assumed to not have dup IDs
    for (key, val) in count_dvid.items():
        if val != 1:
            log.error(f"found a dup dvid { key }: { val } appearances")
    for (key, val) in count_vid.items():
        if val != 1:
            log.error(f"found a dup vid { key }: { val } appearances")

    categories = [ 'R', 'B', 'D', 'P-Van' ]

    print(f"")
    print(f"Report run: { TIMESTAMP }, DR { dr_config.dr_id }")
    print(f"")
    print(f"Vehicle Counts")
    dformat = "%13s %-50s %7d %7d"
    sformat = "%13s %-50s %7s %7s"
    print(sformat % ('5266 Line No', 'Category', 'Period', 'To Date'))
    print(dformat % ('47', 'ERVs', count_category_active.get('A', 0), count_category_all.get('A', 0)))
    print(dformat % ('48', 'Red Cross Vehicles', count_category_active.get('C', 0), count_category_all.get('C', 0)))
    print(f"")
    print(dformat % ('51', 'Passenger Rental Vehicles', count_category_active.get('R', 0), count_category_all.get('R', 0)))
    print(dformat % ('52', 'Passenger Vans', count_category_active.get('P-Van', 0), count_category_all.get('P-Van', 0)))
    print(dformat % ('53', 'Non-Passenger Rental Vehicles/Vans', count_category_active.get('D', 0), count_category_all.get('D', 0)))
    print(dformat % ('54', 'Box Trucks', count_category_active.get('B', 0), count_category_all.get('B', 0)))
    print(dformat % ('55', '  Total Rental Vehicles',
        sumcategory(count_category_active, categories),
        sumcategory(count_category_all, categories)))

    print(f"")
    print(f"")
    print(f"Rental Vehicle Costs")
    dformat = "%6s %-31s %6d %10d"
    sformat = "%6s %-31s %6s %10s"

    total_allpsc_count = 0
    total_allpsc_cost = 0

    cost_table = {
            'R': 45,
            'B': 75,
            'P-Van': 106,
            'D': 70,
            }

    def print_psc(psc, show_details=False):
        nonlocal total_allpsc_count
        nonlocal total_allpsc_cost

        #if psc == 21 or psc == 29:
        #    detailed = True
        #else:
        #    detailed = False


        rental_count = count_rental_psc.get(psc, 0)
        rental_cost = rental_count * cost_table['R']

        boxtruck_count = count_boxtruck_psc.get(psc, 0)
        boxtruck_cost = boxtruck_count * cost_table['B']

        van_count = count_cargovan_psc.get(psc, 0)
        van_cost = van_count * cost_table['D']

        pvan_count = count_passengervan_psc.get(psc, 0)
        pvan_cost = pvan_count * cost_table['P-Van']

        total_count = pvan_count + van_count + boxtruck_count + rental_count
        total_cost = pvan_cost + van_cost + boxtruck_cost + rental_cost

        detailed =  pvan_count > 0 or van_count > 0 or boxtruck_count > 0

        total_allpsc_count += total_count
        total_allpsc_cost += total_cost

        if detailed and show_details:
            print(dformat % (psc, 'Passenger Vehicles (Cat R)', rental_count, rental_cost))
            print(dformat % ('', 'Box Truck (Cat B)', boxtruck_count, boxtruck_cost))
            print(dformat % ('', 'Passenger Van', pvan_count, pvan_cost))
            print(dformat % ('', 'Non-Passenger/Cargo Van (Cat D)', van_count, van_cost))
            print(dformat % ('', f'   Total Cost PSC { psc }', total_count, total_cost))
            print(f"")
        elif show_details == False:
            print(dformat % (psc, 'All Vehicles', total_count, total_cost))


    print(sformat % ( "PSC", "Description", "Veh Count", "Daily Cost"))

    for psc in itertools.chain(range(21, 30), [ 80 ]):
        print_psc(psc)

    print(f"")
    print(dformat % ('', 'Total All Rentals', total_allpsc_count, total_allpsc_cost))

    print(f"")
    print(f"PSC Details")
    for psc in itertools.chain(range(21, 30), [ 80 ]):
        print_psc(psc, show_details = True)

    print(f"")
    print(f"Cost Model:")
    print(f"    Passenger Vehicles (Cat R): { cost_table['R'] }")
    print(f"    Box Trucks (Cat B): { cost_table['B'] }")
    print(f"    Non Passenger Rentals (Cat D): { cost_table['D'] }")
    print(f"    10-14 person Passenger Vans: { cost_table['P-Van'] }")



def parse_args():
    parser = argparse.ArgumentParser(
            description="tools to support Disaster Transportation Tools reporting",
            allow_abbrev=False)
    parser.add_argument("--debug", help="turn on debugging output", action="store_true")
    parser.add_argument("-s", "--store", help="Store file on server", action="store_true")
    parser.add_argument("--do-avis", help="Generate an Avis match report", action="store_true")
    parser.add_argument("--do-group", help="Generate a Group Vehicle report", action="store_true")
    parser.add_argument("--do-vehicles", help="Generate a Vehicle backup", action="store_true")
    parser.add_argument("--do-car", help="Generate vehicle status messages", action="store_true")
    parser.add_argument("--do-no-car", help="Generate vehicle status messages", action="store_true")
    parser.add_argument("--do-dtr", help="generate daily transportation report", action="store_true")
    parser.add_argument("--send", help="Send messages to the actual intended recipients", action="store_true")
    parser.add_argument("--save", help="Keep a copy of the generated report", action="store_true")
    parser.add_argument("--test-send", help="Add the test email account to message recipients", action="store_true")
    parser.add_argument("--mail-limit", help="max number of emails to send (default: 5)", nargs="?", const=5, type=int)

    parser.add_argument("--dr-id", help="the name of the DR (like 155-22)", required=True, action="append")

    group = parser.add_mutually_exclusive_group(required=False)
    group.add_argument("--save-input", help="Save a copy of server inputs", action="store_true")
    group.add_argument("--cached-input", help="Use cached server input", action="store_true")

    args = parser.parse_args()

    if not args.do_avis and not args.do_group and not args.do_car and not args.do_no_car and not args.do_dtr:
        log.error("At least one of do-avis, do-group, do-car, do-no-car, or do-dtr must be specified")
        sys.exit(1)

    return args


if __name__ == "__main__":
    neil_tools.init_logging(__name__)
    log = logging.getLogger(__name__)
    main()
else:
    log = logging.getLogger(__name__)

